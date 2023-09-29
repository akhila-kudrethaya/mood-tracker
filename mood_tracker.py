import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import date

# Function to create a new Excel file or load an existing one
def create_or_load_excel(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Mood Tracker"
        sheet["A1"] = "Date"
        sheet["B1"] = "Mood"
        workbook.save(filename)
    return workbook

# Function to record mood for the day
def record_mood(workbook):
    sheet = workbook.active
    today = date.today()
    today_str = today.strftime("%Y-%m-%d")
    existing_dates = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)]

    # Check if mood for today has already been recorded
    if today_str in existing_dates:
        print(f"Mood for {today_str} has already been recorded.")
        return

    print(f"Recording mood for {today_str}:")
    mood = input("Enter your mood for today (Sad, Neutral, OK, Happy, or Very Happy): ").strip().capitalize()
    if mood not in ["Sad", "Neutral", "OK", "Happy", "Very Happy"]:
        print("Invalid mood. Please enter a valid mood.")
        return

    sheet.append([today_str, mood])
    workbook.save("mood_tracker.xlsx")
    print(f"Mood for {today_str} recorded successfully!")


# Function to create a summary of moods month-wise
def create_monthly_summary(workbook):
    sheet = workbook.active
    summary_workbook = Workbook()
    summary_sheet = summary_workbook.active
    summary_sheet.title = "Monthly Summary"
    summary_sheet["A1"] = "Month"
    summary_sheet["B1"] = "Sad"
    summary_sheet["C1"] = "Neutral"
    summary_sheet["D1"] = "OK"
    summary_sheet["E1"] = "Happy"
    summary_sheet["F1"] = "Very Happy"

    # Create a dictionary to store mood counts for each month
    monthly_mood_counts = {}

    for row in range(2, sheet.max_row + 1):
        date_value = sheet.cell(row=row, column=1).value
        mood = sheet.cell(row=row, column=2).value
        year, month, _ = map(int, date_value.split("-"))
        month_str = f"{year}-{month:02}"

        if month_str not in monthly_mood_counts:
            monthly_mood_counts[month_str] = {"Sad": 0, "Neutral": 0, "OK": 0, "Happy": 0, "Very Happy": 0}

        monthly_mood_counts[month_str][mood] += 1

    # Populate the summary sheet from the dictionary
    for row, (month, mood_counts) in enumerate(monthly_mood_counts.items(), start=2):
        summary_sheet.cell(row=row, column=1, value=month)
        for col, mood in enumerate(["Sad", "Neutral", "OK", "Happy", "Very Happy"], start=2):
            summary_sheet.cell(row=row, column=col, value=mood_counts[mood])

    summary_workbook.save("mood_summary.xlsx")
    print("Monthly summary created successfully!")


if __name__ == "__main__":
    excel_file = "mood_tracker.xlsx"
    workbook = create_or_load_excel(excel_file)
    record_mood(workbook)
    create_monthly_summary(workbook)
